# stats_processing.py
import openpyxl
from collections import defaultdict
import pandas as pd

def process_stats(sheet):
    # Debug - print entire sheet contents
    print("\nDEBUG - Zawartość zakładki 'suma':")
    print("-" * 50)
    
    # Find the dimensions of the used range
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # Print column letters for reference
    col_headers = "    " # indent for row numbers
    for col in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        col_headers += f"{col_letter:^10}"
    print(col_headers)
    print("-" * (max_col * 10 + 4))
    
    # Print each row with row number
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        row_str = f"{row_idx:3d}|"
        for value in row:
            if value is None:
                value = ""
            row_str += f"{str(value):^10}"
        print(row_str)
    
    print("-" * 50)
    
    # Słownik na statystyki graczy
    player_stats = defaultdict(lambda: defaultdict(int))
    
    # Znajdź wiersz z nazwami graczy (wiersz 3)
    players = []
    for row in sheet.iter_rows(min_row=3, max_row=3, values_only=True):
        players = [name.lower() for name in row[3:] if name]  # Start from column D (index 3)
        print(f"\nDEBUG - Znalezieni gracze: {players}")
        break
    
    if not players:
        print("\nDEBUG - Nie znaleziono graczy!")
        return {}

    # Mapowanie kategorii i akcji
    stat_mappings = {
        '!1': 'good receive',
        '!2': 'ok receive',
        '!3': 'bad receive',
        '!4': 'enemy ace',
        '@1': 'ace',           # Asy serwisowe
        '@2': 'serve net',     # Zagrywka w siatkę
        '@3': 'serve out',     # Zagrywka na aut
        '#1': 'position error',
        '#2': 'stance error',
        '#3': 'free ball',
        '#4': 'lost point',
        '$1': 'perfect set',
        '$2': 'good set',
        '$3': 'playable set',
        '$4': 'bad set',
        '%1': 'dig',
        '%2': 'block',
        '^1': 'attack',
        '^2': 'out',
        '^3': 'attack net'
    }

    # Przetwarzanie statystyk
    for row in sheet.iter_rows(min_row=4, max_row=25, values_only=True):
        if not row[1]:  # Pomijamy puste wiersze
            continue
            
        action_code = row[1]  # Kolumna B zawiera kody akcji
        if action_code in stat_mappings:
            action = stat_mappings[action_code]
            
            # Iteruj przez wartości dla każdego gracza (kolumny D-K)
            for i, value in enumerate(row[3:]):
                if i < len(players) and isinstance(value, (int, float)):
                    player_stats[players[i]][action] = value

        # Obsługa sum w wierszach 24-25
        elif row[0] in ['scored', 'lost']:
            action = f"{row[0]} points"
            for i, value in enumerate(row[3:]):
                if i < len(players) and isinstance(value, (int, float)):
                    player_stats[players[i]][action] = value

    # Debug print - końcowe statystyki
    print("\nDEBUG - Końcowe statystyki:")
    for player, stats in player_stats.items():
        print(f"\n{player}:")
        for action, value in stats.items():
            print(f"  {action}: {value}")
    
    return player_stats

def merge_stats(all_stats):
    merged_stats = defaultdict(lambda: defaultdict(int))
    for stats in all_stats:
        for player, player_stats in stats.items():
            player = player.lower()  # Konwertujemy na małe litery
            for action, value in player_stats.items():
                if isinstance(value, (int, float)):  # Upewniamy się, że wartość jest liczbą
                    merged_stats[player][action] += value
    return merged_stats

def format_player_stats(player_stats, is_summary=False):
    if not player_stats:
        return "Nie znaleziono statystyk graczy\n"
    
    title = "\nPODSUMOWANIE WSZYSTKICH MECZÓW:\n" if is_summary else "\nStatystyki graczy:\n"
    output = title
    output += "-" * 50 + "\n"
    
    for player, stats in player_stats.items():
        output += f"\n{player.upper()}:\n"
        
        # Statystyki ataku
        attacks = stats.get('attack', 0)  # punkty z ataku
        attack_errors = stats.get('attack net', 0) + stats.get('out', 0)  # tylko net i out
        total_attacks = attacks + attack_errors
        if total_attacks > 0:
            attack_efficiency = (attacks / total_attacks) * 100
        else:
            attack_efficiency = 0
            
        output += f"ATAK:\n"
        output += f"  Punkty: {attacks}\n"
        output += f"  Błędy: {attack_errors}\n"
        output += f"  Skuteczność: {attack_efficiency:.1f}%\n"
        
        # Statystyki zagrywki
        aces = stats.get('ace', 0)  # Używamy 'ace' zamiast '@1'
        serve_errors = stats.get('serve net', 0) + stats.get('serve out', 0)
        total_serves = aces + serve_errors
        if total_serves > 0:
            serve_efficiency = (aces / total_serves) * 100
        else:
            serve_efficiency = 0
            
        output += f"\nZAGRYWKA:\n"
        output += f"  Asy: {aces}\n"
        output += f"  Błędy: {serve_errors}\n"
        output += f"  Skuteczność: {serve_efficiency:.1f}%\n"
        
        # Statystyki przyjęcia
        good_receives = stats.get('good receive', 0)
        ok_receives = stats.get('ok receive', 0)
        bad_receives = stats.get('bad receive', 0)
        enemy_aces = stats.get('enemy ace', 0)
        total_receives = good_receives + ok_receives + bad_receives + enemy_aces
        
        if total_receives > 0:
            # Nowa formuła: tylko dobre i średnie jako udane
            receive_efficiency = ((good_receives + ok_receives) / total_receives) * 100
        else:
            receive_efficiency = 0
        
        output += f"\nPRZYJĘCIE:\n"
        output += f"  Idealne: {good_receives}\n"
        output += f"  Dość dobre: {ok_receives}\n"
        output += f"  Słabe: {bad_receives}\n"
        output += f"  Asy przeciwnika: {enemy_aces}\n"
        output += f"  Łącznie przyjęć: {total_receives}\n"
        output += f"  Skuteczność: {receive_efficiency:.1f}%\n"
        
        # Pozostałe statystyki
        blocks = stats.get('block', 0)
        free_balls = stats.get('free ball', 0)
        
        output += f"\nPOZOSTAŁE:\n"
        output += f"  Bloki: {blocks}\n"
        output += f"  Free ball: {free_balls}\n"
        output += "-" * 30 + "\n"
    
    return output

def create_player_summary_df(player_stats):
    """Tworzy DataFrame z podsumowaniem statystyk gracza"""
    data = []
    for player, stats in player_stats.items():
        # Atak
        attacks = stats.get('attack', 0)
        attack_nets = stats.get('attack net', 0)
        outs = stats.get('out', 0)
        position_errors = stats.get('position error', 0)
        stance_errors = stats.get('stance error', 0)
        
        # Zagrywka
        aces = stats.get('ace', 0)
        serve_nets = stats.get('serve net', 0)
        serve_outs = stats.get('serve out', 0)
        
        # Przyjęcie
        good_receives = stats.get('good receive', 0)
        ok_receives = stats.get('ok receive', 0)
        bad_receives = stats.get('bad receive', 0)
        enemy_aces = stats.get('enemy ace', 0)
        
        # Rozegranie
        perfect_sets = stats.get('perfect set', 0)
        good_sets = stats.get('good set', 0)
        playable_sets = stats.get('playable set', 0)
        bad_sets = stats.get('bad set', 0)
        
        # Obrona i inne
        digs = stats.get('dig', 0)
        blocks = stats.get('block', 0)
        free_balls = stats.get('free ball', 0)
        lost_points = stats.get('lost point', 0)  # Punkty stracone z #4
        
        # Sumy punktów
        scored_points = stats.get('scored points', 0)  # Z wiersza 24
        lost_points_total = stats.get('lost points', 0)  # Z wiersza 25
        
        data.append({
            'Gracz': player,
            'Punkty z ataku': attacks,
            'Atak w siatkę': attack_nets,
            'Aut': outs,
            'Błędy pozycji': position_errors,
            'Błędy postawy': stance_errors,
            'Asy': aces,
            'Zagrywka w siatkę': serve_nets,
            'Zagrywka na aut': serve_outs,
            'Dobre przyjęcia': good_receives,
            'Średnie przyjęcia': ok_receives,
            'Złe przyjęcia': bad_receives,
            'Przyjęte asy': enemy_aces,
            'Perfect set': perfect_sets,
            'Good set': good_sets,
            'Playable set': playable_sets,
            'Bad set': bad_sets,
            'Dig': digs,
            'Bloki': blocks,
            'Free ball': free_balls,
            'Stracone punkty (inne)': lost_points,  # Z #4
            'Suma zdobytych': scored_points,  # Z wiersza 24
            'Suma straconych': lost_points_total  # Z wiersza 25
        })
    
    return pd.DataFrame(data)
